import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def save_to_json(results, output_path="output/cvs_data.json"):
    """
    Sauvegarde tous les CVs en JSON
    """
    os.makedirs("output", exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"✅ JSON sauvegardé : {output_path}")


def format_scores_list(items):
    """
    Formate une liste [{"domaine": ..., "score": ...}] en chaîne lisible pour Excel.
    Ex: [{"domaine": "Dev Web", "score": 85}] -> "Dev Web: 85, IA: 70"

    Gère aussi l'ancien format (dict) au cas où un vieux JSON traînerait encore,
    pour ne pas planter dessus.
    """
    if not items:
        return ""

    # Nouveau format : liste de dicts {"domaine": ..., "score": ...}
    if isinstance(items, list):
        parts = []
        for item in items:
            if isinstance(item, dict) and "domaine" in item:
                parts.append(f"{item['domaine']}: {item.get('score', '')}")
        return ", ".join(parts)

    # Ancien format (fallback de sécurité) : dict {domaine: score}
    if isinstance(items, dict):
        return ", ".join(f"{k}: {v}" for k, v in items.items())

    return ""


def format_dict_scores(d):
    """
    Formate un dict à clés fixes (ex: details_score_qualite) en chaîne lisible.
    Ce champ-là reste un dict classique (diplome, certifications, projets...),
    pas de risque d'explosion puisque les clés sont toujours les mêmes.
    """
    if not d or not isinstance(d, dict):
        return ""
    return ", ".join(f"{k}: {v}" for k, v in d.items())


def save_to_excel(results, output_path="output/cvs_data.xlsx"):
    """
    Sauvegarde tous les CVs en Excel, avec les scores de qualité
    et de pertinence par domaine.
    """
    os.makedirs("output", exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CVs"

    # Style entête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F75B6")

    # Colonnes
    headers = [
        "Fichier", "Nom", "Email", "Telephone", "LinkedIn", "Localisation",
        "Categorie principale",
        "Langages", "Frameworks", "Bases de données", "Outils DevOps",
        "Technologies", "Projets", "Diplômes", "Certifications", "Langues",
        "Score qualité (/100)", "Score qualité (/10)", "Détail score qualité",
        "Domaines & pertinence (LLM)", "Domaines pondérés (score final)",
        "Années d'expérience (pondérées)", "Alertes parcours",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        width = 40 if header in [
            "Détail score qualité", "Domaines & pertinence (LLM)",
            "Domaines pondérés (score final)", "Technologies", "Projets",
            "Alertes parcours",
        ] else 25
        ws.column_dimensions[cell.column_letter].width = width

    # Écrire les données
    for row, result in enumerate(results, 2):
        data = result.get("data") or {}

        def safe_join(field):
            value = data.get(field) or []
            return ", ".join(value) if isinstance(value, list) else ""

        ws.cell(row=row, column=1, value=result["filename"])
        ws.cell(row=row, column=2, value=data.get("nom", ""))
        ws.cell(row=row, column=3, value=data.get("email", ""))
        ws.cell(row=row, column=4, value=data.get("telephone", ""))
        ws.cell(row=row, column=5, value=data.get("linkedin", ""))
        ws.cell(row=row, column=6, value=data.get("localisation", ""))
        ws.cell(row=row, column=7, value=data.get("categorie_principale", ""))
        ws.cell(row=row, column=8, value=safe_join("langages"))
        ws.cell(row=row, column=9, value=safe_join("frameworks"))
        ws.cell(row=row, column=10, value=safe_join("bases_de_donnees"))
        ws.cell(row=row, column=11, value=safe_join("outils_devops"))
        ws.cell(row=row, column=12, value=safe_join("technologies"))
        ws.cell(row=row, column=13, value=safe_join("projets"))
        ws.cell(row=row, column=14, value=safe_join("diplomes"))
        ws.cell(row=row, column=15, value=safe_join("certifications"))
        ws.cell(row=row, column=16, value=safe_join("langues"))

        # --- Scores de qualité ---
        ws.cell(row=row, column=17, value=data.get("score_qualite_globale", ""))
        ws.cell(row=row, column=18, value=data.get("score_qualite_globale_sur_10", ""))
        ws.cell(row=row, column=19, value=format_dict_scores(data.get("details_score_qualite")))
        # Ces deux-là sont maintenant des LISTES (nouveau format nested)
        ws.cell(row=row, column=20, value=format_scores_list(data.get("scores_categories")))
        ws.cell(row=row, column=21, value=format_scores_list(data.get("scores_categories_ponderes")))

        # --- Expérience pondérée & alertes de parcours ---
        ws.cell(row=row, column=22, value=data.get("annees_experience", ""))
        alertes = data.get("alertes_parcours") or []
        ws.cell(row=row, column=23, value=" | ".join(alertes) if alertes else "")

        if not data and result.get("error"):
            ws.cell(row=row, column=2, value=f"❌ ÉCHEC : {result['error']}")

    wb.save(output_path)
    print(f"✅ Excel sauvegardé : {output_path}")


# TEST
if __name__ == "__main__":
    from cv_reader import read_all_cvs
    from cv_extractor import extract_all_cvs
    from cv_deduplication import detect_duplicates
    from cv_cache import CVCache, file_sha256
    import os, json

    # Load raw CV entries (filename + raw text)
    raw_entries = read_all_cvs("cvs/")

    # Initialise cache (Postgres if env vars set, otherwise JSON fallback)
    cache = CVCache()
    processed_entries = []

    for raw in raw_entries:
        pdf_path = os.path.join("cvs", raw["filename"])
        h = file_sha256(pdf_path)
        cached = cache.get_by_hash(h)
        if cached:
            # Cache hit – reuse stored extraction result
            processed_entries.append({
                "filename": raw["filename"],
                "data": cached.get("data"),
                "text": cached.get("text", ""),
                "hash": h,
            })
            continue
        # Cache miss – run extraction and store result
        extracted = extract_all_cvs([raw])
        if not extracted:
            continue
        entry = extracted[0]
        entry["hash"] = h
        entry["source_path"] = pdf_path
        # Insert into cache for future runs
        cache.insert({
            "hash": h,
            "email": (entry.get("data", {}) or {}).get("email", "").strip().lower(),
            "phone": (entry.get("data", {}) or {}).get("telephone", "").strip(),
            "data": entry.get("data"),
            "text": entry.get("text", ""),
            "source_path": pdf_path,
        })
        processed_entries.append(entry)

    # Deduplication (levels 1‑3)
    unique_results, duplicate_report = detect_duplicates(processed_entries)

    # Save duplicate report
    if duplicate_report:
        os.makedirs("output", exist_ok=True)
        dup_path = os.path.join("output", "duplicates.json")
        with open(dup_path, "w", encoding="utf-8") as f:
            json.dump(duplicate_report, f, ensure_ascii=False, indent=2)
        print(f"✅ Rapport des doublons sauvegardé : {dup_path}")

    # Persist unique results
    save_to_json(unique_results)
    save_to_excel(unique_results)

    print("\n🎉 Terminé ! Fichiers dans le dossier output/")